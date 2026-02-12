import warnings
warnings.simplefilter(action='ignore')

import os
import oracledb
import pandas as pd
from openpyxl import load_workbook
from copy import copy
from queries import load_queries
from datetime import datetime
import timeit


def connect_to_DB (username,password,hostname):
    """ Returns a connection and cursor to Oracle database"""
    try:
        connection = oracledb.connect(user=username, password=password, dsn=hostname)
        cursor = connection.cursor()
        print  ("....Successffuly connected to the database")
    except:
        raise Exception('....Connection failed! Please check your login parameters')

    return connection, cursor


def insert_master_index(report_file, index_file):
    """
    Copies the 'Master Index' sheet (with all styling) from the index file
    into the report workbook as the first sheet '0-Master Index'.
    """
    if not os.path.exists(index_file):
        print(f'    WARNING: Index file not found at {index_file}. Skipping Master Index.')
        return

    print('     Inserting styled Master Index into report.')

    src_wb = load_workbook(index_file)
    src_ws = src_wb['Master Index']

    dst_wb = load_workbook(report_file)
    dst_ws = dst_wb.create_sheet('0-Master Index', 0)  # insert at position 0

    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = dim.width
        dst_ws.column_dimensions[col_letter].hidden = dim.hidden

    # Copy row heights
    for row_num, dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = dim.height

    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # Copy cell values, styles, and formatting
    for row in src_ws.iter_rows():
        for cell in row:
            new_cell = dst_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = cell.number_format
                new_cell.protection = copy(cell.protection)

    # Copy auto-filter if present
    if src_ws.auto_filter.ref:
        dst_ws.auto_filter.ref = src_ws.auto_filter.ref

    # Copy freeze panes
    if src_ws.freeze_panes:
        dst_ws.freeze_panes = src_ws.freeze_panes

    dst_wb.save(report_file)
    src_wb.close()
    dst_wb.close()
    print('     Master Index inserted successfully.')


def generate_report (workspace, df_list, sheet_list, filename):
    """ Exports dataframes to multi-tab excel spreadsheet"""
    outfile= os.path.join(workspace, filename + '.xlsx')

    writer = pd.ExcelWriter(outfile, engine='xlsxwriter')
    workbook = writer.book

    # Bold format for "No records" message
    bold_format = workbook.add_format({'bold': True})

    for dataframe, sheet in zip(df_list, sheet_list):
        dataframe = dataframe.reset_index(drop=True)
        dataframe.index = dataframe.index + 1

        # Handle empty dataframes - preserve column names, add bold message
        if dataframe.empty:
            dataframe.to_excel(writer, sheet_name=sheet, index=False, startrow=0, startcol=0)
            worksheet = writer.sheets[sheet]
            worksheet.write(1, 0, 'No records returned for this query', bold_format)
            worksheet.set_column(0, max(dataframe.shape[1] - 1, 0), 20)
        else:
            dataframe.to_excel(writer, sheet_name=sheet, index=False, startrow=0 , startcol=0)

            worksheet = writer.sheets[sheet]

            worksheet.set_column(0, dataframe.shape[1], 20)

            col_names = [{'header': col_name} for col_name in dataframe.columns[1:-1]]
            col_names.insert(0,{'header' : dataframe.columns[0], 'total_string': 'Total'})
            col_names.append ({'header' : dataframe.columns[-1], 'total_function': 'sum'})

            worksheet.add_table(0, 0, dataframe.shape[0]+1, dataframe.shape[1]-1, {
                'total_row': True,
                'columns': col_names})

    writer.close()

    return outfile


if __name__ == "__main__":
    start_t = timeit.default_timer() #start time

    workspace = r"\\spatialfiles.bcgov\work\srm\gss\projects\gr_2026_197_haida_gwaii_Q4_interests"
    report_folder = os.path.join(workspace, 'deliverables')
    index_file = os.path.join(report_folder, 'Haida_Gwaii_Interests_Report_Index.xlsx')

    print ('\nConnecting to BCGW.')
    hostname = 'bcgw.bcgov/idwprod1.bcgov'

    bcgw_user = os.getenv('bcgw_user')
    bcgw_pwd = os.getenv('bcgw_pwd')

    connection, cursor = connect_to_DB (bcgw_user,bcgw_pwd,hostname)

    print ('\nRunning queries.')
    sql = load_queries()
    df_list = []
    sheet_list = []
    total_queries = len(sql)
    query_count = 0
    for k, v in sql.items():
        query_count += 1
        print (f'***working on: query {query_count} of {total_queries}: {k}***')
        df = pd.read_sql(v, con=connection)
        print (f'   nbr of records returned: {len(df)}')
        df_list.append(df)
        if len(k) > 31:
            k = k[:31]
        sheet_list.append(k)
    
    print ('\nExporting report to Excel.')
    today = datetime.today().strftime('%Y-%m-%d')
    report_file = generate_report(
        report_folder, df_list, sheet_list, 
        f'{today}_Haida_Gwaii_Interests_Report'
    )
    
    # Insert styled Master Index as first sheet (post-processing)
    print ('\nAdding Master Index.')
    insert_master_index(report_file, index_file)
    
    finish_t = timeit.default_timer() #finish time
    t_sec = round(finish_t-start_t)
    mins = int (t_sec/60)
    secs = int (t_sec%60)
    print ('\nProcessing Completed in {} minutes and {} seconds'.format (mins,secs))